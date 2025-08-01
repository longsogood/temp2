import os
import base64
import io
import openpyxl
from openpyxl.utils import get_column_letter

def encode_s3_path(path):
    """
    Mã hóa đường dẫn S3: thay thế dấu cách bằng %20 và dấu / bằng %2F, giữ nguyên các ký tự khác.
    """
    return path.replace(' ', '%20').replace('/', '%2F')

def format_markdown_cell(text):
    if not text:
        return ""
    text = text.replace("|", "\\|")
    text = text.replace("\n", " ")
    return text.strip()

def process_xlsx(file_name: str, file_content: bytes):
    """
    Đọc file Excel (nhận vào bytes) và trả về nội dung markdown cho từng sheet, mỗi sheet có dạng:
    ## Sheet [tên sheet]
    [Nội dung sheet]
    """
    def format_cell(text):
        if text is None:
            return ""
        return str(text).replace("|", "\\|").replace("\n", "<br>")

    try:
        # file_content là bytes
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        output = []
        for sheet_name in wb.sheetnames:
            sheet_name_clean = sheet_name.strip()
            sheet = wb[sheet_name]
            image_map = {}

            # Xử lý ảnh (nếu có)
            for image in getattr(sheet, '_images', []):
                try:
                    anchor = image.anchor
                    if hasattr(anchor, "_from"):
                        from_col = anchor._from.col
                        from_row = anchor._from.row
                        coord = f"{get_column_letter(from_col + 1)}{from_row + 1}"
                    else:
                        coord = getattr(anchor, 'cell', None) or str(anchor) or "UNKNOWN"
                except Exception:
                    coord = "UNKNOWN"
                image_map[coord] = image

            rows = list(sheet.iter_rows(values_only=False))

            # Tìm cột cuối cùng có dữ liệu hoặc hình ảnh
            max_col = 0
            for coord in image_map.keys():
                if coord == "UNKNOWN":
                    continue
                col_letter = ''.join(filter(str.isalpha, coord))
                try:
                    col_index = openpyxl.utils.column_index_from_string(col_letter)
                    max_col = max(max_col, col_index)
                except Exception:
                    continue

            for row in rows:
                for col in range(1, sheet.max_column + 3):
                    cell = sheet.cell(row=row[0].row, column=col)
                    if cell.value is not None and str(cell.value).strip():
                        max_col = max(max_col, col)

            # Tạo header hàng đầu: 1 ô trống cho cột số dòng, rồi A, B, C,...
            header_row = [""]
            for col_idx in range(1, max_col + 1):
                header_row.append(get_column_letter(col_idx))

            # Tạo header dữ liệu chính (header dòng 1 trong Excel)
            data_header = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col_idx)
                data_header.append(format_cell(cell.value) if cell.value is not None else "")

            # Tạo bảng markdown
            all_rows = []
            all_rows.append(header_row)
            all_rows.append([""] + data_header)

            for row_idx, row in enumerate(rows[1:], start=2):
                row_data = [str(row_idx - 1)]
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    coord = cell.coordinate
                    try:
                        if hasattr(cell, 'hyperlink') and cell.hyperlink is not None:
                            link = cell.hyperlink.target
                            display_text = str(cell.value).strip() if cell.value else link
                            text = f"[{display_text}]({link})"
                        else:
                            text = str(cell.value) if cell.value else ""

                        if coord in image_map:
                            text += f"\nTham khảo ảnh chụp màn hình tại ô {coord} trong file `{file_name}`, sheet `{sheet_name_clean}`.\n"
                        elif "UNKNOWN" in image_map:
                            text += f"\nTham khảo ảnh chụp màn hình trong file `{file_name}`, sheet `{sheet_name_clean}`.\n"

                        row_data.append(format_cell(text))
                    except Exception:
                        row_data.append("")
                all_rows.append(row_data)

            # Tạo markdown table cho sheet này
            markdown_content = []
            markdown_content.append("| " + " | ".join(all_rows[0]) + " |")
            markdown_content.append("|" + "|".join(["---"] * len(all_rows[0])) + "|")
            for row in all_rows[1:]:
                markdown_content.append("| " + " | ".join(row) + " |")

            # Thêm vào output với tiêu đề sheet
            output.append(f"## Sheet {sheet_name_clean}\n" + "\n".join(markdown_content))

        wb.close()
        return "\n\n".join(output)
    except Exception as e:
        # Có thể log lỗi ở đây nếu cần
        raise